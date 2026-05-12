import csv
import io
from datetime import date
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from accounts.models import CompanyProfile
from analytics.services import (
    getViewsVsApplicationsPerJob, getAverageTimeToClose,
    getCandidatesByStatus, is_profile_complete, get_demanded_skills,
    get_export_applications, get_export_offers,
)


@login_required
def company_metrics_page(request, company_id):
    if request.user.role != 'company':
        return redirect('home')
    profile = CompanyProfile.objects.get(user=request.user)
    if profile.id != company_id:
        return redirect('home')
    return render(request, 'analytics/company_metrics.html', {'company': profile})


@login_required
def company_metrics_api(request, id):
    company_id = id
    if request.user.role != 'company':
        return JsonResponse({'error': 'Forbidden'}, status=403)

    profile = CompanyProfile.objects.get(user=request.user)
    if profile.id != company_id:
        return JsonResponse({'error': 'Not found'}, status=404)

    filters = {}
    month = request.GET.get('month')
    year = request.GET.get('year')
    try:
        if month:
            filters['month'] = int(month)
        if year:
            filters['year'] = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid month or year'}, status=400)

    metrics = getViewsVsApplicationsPerJob(company_id, filters)
    average_time_to_close = getAverageTimeToClose(company_id, filters)
    candidates_by_status = getCandidatesByStatus(company_id, filters)

    total_views = sum(item['views_count'] for item in metrics)
    total_applications = sum(item['applications_count'] for item in metrics)

    return JsonResponse({
        'views_vs_applications': metrics,
        'average_time_to_close_days': average_time_to_close,
        'candidates_by_status': candidates_by_status,
        'totals': {
            'total_views': total_views,
            'total_applications': total_applications,
            'average_time_to_close_days': average_time_to_close,
        }
    })


@login_required
def profile_status(request):
    status = is_profile_complete(request.user)
    return JsonResponse(status)


@login_required
def demanded_skills_api(request):
    if request.user.role != 'student':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    career_id = request.GET.get('career_id')
    if career_id is not None:
        try:
            career_id = int(career_id)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid career_id'}, status=400)

    result = get_demanded_skills(career_id=career_id)
    return JsonResponse(result)


# ── US-20: Data Export ────────────────────────────────────────────────────────

def _get_export_params(request):
    """Parse and return common export filter params."""
    date_from = request.GET.get('date_from') or None
    date_to   = request.GET.get('date_to')   or None
    status    = request.GET.get('status')    or None
    fmt       = request.GET.get('format', 'csv').lower()
    if fmt not in ('csv', 'xlsx'):
        fmt = 'csv'
    return date_from, date_to, status, fmt


@login_required
def export_page(request):
    if request.user.role != 'company':
        return redirect('home')
    return render(request, 'analytics/export.html')


@login_required
def export_check(request):
    """Returns record count for preview warning without generating file."""
    if request.user.role != 'company':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    profile     = CompanyProfile.objects.get(user=request.user)
    export_type = request.GET.get('type', 'applications')
    date_from, date_to, status, _ = _get_export_params(request)

    if export_type == 'offers':
        count = get_export_offers(profile, date_from, date_to, status).count()
    else:
        count = get_export_applications(profile, date_from, date_to, status).count()

    return JsonResponse({'count': count})


@login_required
def export_applications(request):
    if request.user.role != 'company':
        return redirect('home')

    profile = CompanyProfile.objects.get(user=request.user)
    date_from, date_to, status, fmt = _get_export_params(request)
    qs = get_export_applications(profile, date_from, date_to, status)

    headers = ['Student name', 'National ID', 'Career(s)', 'University',
               'Company', 'Offer title', 'Applied at', 'Status']

    def row(app):
        careers = ', '.join(app.student.careers.values_list('name', flat=True))
        return [
            app.student.user.get_full_name() or app.student.user.username,
            app.student.cedula,
            careers,
            app.student.university,
            app.offer.company.company_name,
            app.offer.title,
            app.applied_at.strftime('%Y-%m-%d %H:%M'),
            app.get_status_display(),
        ]

    filename = f"applications_{date.today()}"
    return _build_response(qs, headers, row, filename, fmt)


@login_required
def export_offers(request):
    if request.user.role != 'company':
        return redirect('home')

    profile = CompanyProfile.objects.get(user=request.user)
    date_from, date_to, status, fmt = _get_export_params(request)
    qs = get_export_offers(profile, date_from, date_to, status)

    headers = ['Title', 'Company', 'Required skills', 'Modality',
               'Location', 'Salary', 'Created at', 'Status']

    def row(offer):
        return [
            offer.title,
            offer.company.company_name,
            offer.desired_skills,
            offer.get_modality_display(),
            offer.location,
            str(offer.salary) if offer.salary else '',
            offer.created_at.strftime('%Y-%m-%d'),
            offer.get_status_display(),
        ]

    filename = f"offers_{date.today()}"
    return _build_response(qs, headers, row, filename, fmt)


def _build_response(queryset, headers, row_fn, filename, fmt):
    """Build CSV or XLSX HttpResponse from a queryset."""
    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        # Style header row
        header_fill = PatternFill('solid', fgColor='111827')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for obj in queryset:
            ws.append(row_fn(obj))
        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    # CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for obj in queryset:
        writer.writerow(row_fn(obj))
    return response
